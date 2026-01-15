#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script để so sánh kết quả facet giữa 3 Solr containers
- Lấy 100 documents đầu tiên từ Solr
- Query từng ID trên cả 3 containers
- So sánh kết quả facet giữa các containers

Cách sử dụng:
    python compare_facet_results.py [num_docs] [source_port]
    
Tham số:
    num_docs: Số documents để lấy (mặc định: 100)
    source_port: Port của Solr để lấy danh sách documents (mặc định: 8983)
"""

import requests
import json
import sys
from urllib.parse import urlencode
from collections import defaultdict
import time
from datetime import datetime

# Thử import openpyxl, nếu không có thì sẽ báo lỗi khi cần
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Tham số từ command line
NUM_DOCS = int(sys.argv[1]) if len(sys.argv) > 1 else 1000
SOURCE_PORT = int(sys.argv[2]) if len(sys.argv) > 2 else 8983

# Query parameters cho facet
FACET_PARAMS = {
    "q": "*:*",
    "facet": "true",
    "facet.field": "search_text_cloud",
    "facet.sort": "count",
    "rows": "0",
    "wt": "json",
    "indent": "true",
    "facet.limit": "1000",
    "facet.mincount": "1"
}

# Container configurations
CONTAINERS = [
    {
        "name": "solr_8_5_2_1_1",
        "port": 8983,
        "core": "topic_tanvd",
        "version": "Solr 8.5.2 (VnCoreNLP 1.1.1)"
    },
    {
        "name": "solr_8_5_2_1_2",
        "port": 8984,
        "core": "topic_tanvd",
        "version": "Solr 8.5.2 (VnCoreNLP 1.2)"
    },
    {
        "name": "solr_9_11",
        "port": 8985,
        "core": "topic_tanvd_9",
        "version": "Solr 9.11"
    }
]


def get_document_ids(port, core, num_docs):
    """Lấy danh sách ID từ Solr"""
    url = f"http://localhost:{port}/solr/{core}/select"
    params = {
        "q": "*:*",
        "rows": num_docs,
        "fl": "id",
        "wt": "json"
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        ids = [doc["id"] for doc in data.get("response", {}).get("docs", [])]
        return ids
    except Exception as e:
        print(f"❌ ERROR khi lấy document IDs: {str(e)}")
        return []


def get_search_text(port, core, doc_id):
    """Lấy field search_text từ document"""
    url = f"http://localhost:{port}/solr/{core}/select"
    params = {
        "q": f"id:{doc_id}",
        "fl": "search_text",
        "rows": 1,
        "wt": "json"
    }
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        docs = data.get("response", {}).get("docs", [])
        if docs and "search_text" in docs[0]:
            search_text = docs[0]["search_text"]
            # Nếu là list, join lại thành string
            if isinstance(search_text, list):
                return "\n".join(str(item) for item in search_text)
            return str(search_text)
        return ""
    except Exception as e:
        return ""


def get_facet_results(port, core, doc_id):
    """Lấy kết quả facet cho một document ID"""
    url = f"http://localhost:{port}/solr/{core}/select"
    params = FACET_PARAMS.copy()
    params["fq"] = f"id:{doc_id}"
    
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Extract facet terms
        facet_field = data.get("facet_counts", {}).get("facet_fields", {}).get("search_text_cloud", [])
        
        # Convert từ array [term1, count1, term2, count2, ...] thành dict
        facet_dict = {}
        for i in range(0, len(facet_field), 2):
            if i + 1 < len(facet_field):
                term = facet_field[i]
                count = facet_field[i + 1]
                facet_dict[term] = count
        
        return facet_dict, data.get("response", {}).get("numFound", 0)
    except Exception as e:
        print(f"   ❌ ERROR khi query ID {doc_id}: {str(e)}")
        return {}, 0


def compare_facet_results(results_dict):
    """So sánh kết quả facet giữa các containers"""
    comparisons = []
    
    # So sánh từng cặp containers
    for i in range(len(CONTAINERS)):
        for j in range(i + 1, len(CONTAINERS)):
            container1 = CONTAINERS[i]
            container2 = CONTAINERS[j]
            
            version1 = container1["version"]
            version2 = container2["version"]
            
            # Đếm số documents có kết quả giống nhau và khác nhau
            same_count = 0
            diff_count = 0
            only_in_1 = 0
            only_in_2 = 0
            total_terms_diff = 0
            
            for doc_id in results_dict:
                facets1 = results_dict[doc_id].get(version1, {})
                facets2 = results_dict[doc_id].get(version2, {})
                
                if facets1 == facets2:
                    same_count += 1
                else:
                    diff_count += 1
                    
                    # Tính toán sự khác biệt
                    terms1 = set(facets1.keys())
                    terms2 = set(facets2.keys())
                    
                    only_in_1_count = len(terms1 - terms2)
                    only_in_2_count = len(terms2 - terms1)
                    
                    if only_in_1_count > 0:
                        only_in_1 += 1
                    if only_in_2_count > 0:
                        only_in_2 += 1
                    
                    total_terms_diff += abs(len(terms1) - len(terms2))
            
            comparisons.append({
                "container1": version1,
                "container2": version2,
                "same": same_count,
                "different": diff_count,
                "only_in_1": only_in_1,
                "only_in_2": only_in_2,
                "avg_terms_diff": total_terms_diff / diff_count if diff_count > 0 else 0
            })
    
    return comparisons


def export_to_excel(results_dict, search_text_dict, ids, timestamp, logger):
    """Xuất kết quả facet ra file Excel"""
    if not HAS_OPENPYXL:
        logger.log("⚠️  Thư viện openpyxl chưa được cài đặt. Không thể tạo file Excel.")
        logger.log("   Cài đặt bằng lệnh: pip install openpyxl")
        return None
    
    excel_file = f"facet_comparison_results_{timestamp}.xlsx"
    
    logger.log("━" * 70)
    logger.log("Bước 6: Xuất kết quả ra file Excel")
    logger.log("━" * 70)
    logger.log()
    logger.log(f"📊 Đang tạo file Excel: {excel_file}")
    
    # Tạo workbook và worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Facet Comparison"
    
    # Định nghĩa header
    headers = ["Document ID", "search_text", "Solr 8.5.2 (VnCoreNLP 1.1.1)", "Solr 8.5.2 (VnCoreNLP 1.2)", "Solr 9.11"]
    
    # Style cho header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Ghi header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Style cho data cells
    data_alignment = Alignment(vertical="top", wrap_text=True)
    data_font = Font(size=10)
    
    # Ghi dữ liệu từng document
    row_idx = 2
    for doc_id in ids:
        # Cột 1: Document ID
        ws.cell(row=row_idx, column=1, value=doc_id).font = Font(bold=True, size=10)
        
        # Cột 2: search_text
        search_text = search_text_dict.get(doc_id, "")
        cell_search_text = ws.cell(row=row_idx, column=2, value=search_text)
        cell_search_text.font = data_font
        cell_search_text.alignment = data_alignment
        
        # Cột 3-5: Facet results cho từng container
        for col_idx, container in enumerate(CONTAINERS, 3):
            facets = results_dict[doc_id].get(container["version"], {})
            
            if not facets:
                cell_value = "Document không tồn tại"
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = Font(size=10, italic=True, color="808080")
            else:
                # Format facet results: term1 (count1), term2 (count2), ...
                facet_items = []
                for term, count in sorted(facets.items(), key=lambda x: (-x[1], x[0])):  # Sort by count desc, then term
                    facet_items.append(f"{term} ({count})")
                
                cell_value = "\n".join(facet_items)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.font = data_font
                cell.alignment = data_alignment
            
            cell.alignment = data_alignment
        
        row_idx += 1
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 40  # Document ID
    ws.column_dimensions['B'].width = 60  # search_text
    ws.column_dimensions['C'].width = 50  # Solr 8.5.2 (VnCoreNLP 1.1.1)
    ws.column_dimensions['D'].width = 50  # Solr 8.5.2 (VnCoreNLP 1.2)
    ws.column_dimensions['E'].width = 50  # Solr 9.11
    
    # Đặt chiều cao hàng tự động
    for row in ws.iter_rows(min_row=2, max_row=row_idx):
        max_lines = 1
        for cell in row[1:]:  # Bỏ qua cột Document ID
            if cell.value:
                lines = str(cell.value).count('\n') + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = min(max_lines * 15, 300)  # Max 300px
    
    # Đóng băng hàng đầu tiên (header) và cột Document ID
    ws.freeze_panes = 'B2'
    
    # Tạo sheet thống kê
    stats_ws = wb.create_sheet("Statistics")
    
    # Header cho sheet Statistics
    stats_headers = ["Metric", "Value"]
    for col_idx, header in enumerate(stats_headers, 1):
        cell = stats_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Tính toán thống kê
    total_docs = len(ids)
    stats_data = [
        ["Tổng số documents", total_docs],
        ["Solr 8.5.2 (VnCoreNLP 1.1.1)", ""],
        ["  - Documents có facet", sum(1 for doc_id in ids if results_dict[doc_id].get(CONTAINERS[0]["version"], {}))],
        ["Solr 8.5.2 (VnCoreNLP 1.2)", ""],
        ["  - Documents có facet", sum(1 for doc_id in ids if results_dict[doc_id].get(CONTAINERS[1]["version"], {}))],
        ["Solr 9.11", ""],
        ["  - Documents có facet", sum(1 for doc_id in ids if results_dict[doc_id].get(CONTAINERS[2]["version"], {}))],
    ]
    
    # Tính số documents có sự khác biệt
    same_count = 0
    diff_count = 0
    for doc_id in ids:
        facets1 = results_dict[doc_id].get(CONTAINERS[0]["version"], {})
        facets2 = results_dict[doc_id].get(CONTAINERS[1]["version"], {})
        facets3 = results_dict[doc_id].get(CONTAINERS[2]["version"], {})
        
        if facets1 == facets2 == facets3:
            same_count += 1
        else:
            diff_count += 1
    
    stats_data.extend([
        ["", ""],
        ["So sánh", ""],
        ["  - Documents giống nhau (cả 3 containers)", same_count],
        ["  - Documents khác nhau", diff_count],
    ])
    
    # Ghi thống kê
    for row_idx, (metric, value) in enumerate(stats_data, 2):
        stats_ws.cell(row=row_idx, column=1, value=metric).font = data_font
        stats_ws.cell(row=row_idx, column=2, value=value).font = data_font
    
    # Điều chỉnh độ rộng cột cho sheet Statistics
    stats_ws.column_dimensions['A'].width = 50
    stats_ws.column_dimensions['B'].width = 20
    
    # Lưu file
    wb.save(excel_file)
    logger.log(f"✅ Đã tạo file Excel: {excel_file}")
    logger.log()
    
    return excel_file


class Logger:
    """Class để log vừa ra console vừa vào file"""
    def __init__(self, log_file):
        self.log_file = log_file
        self.file = open(log_file, 'w', encoding='utf-8')
    
    def log(self, message='', end='\n'):
        """Ghi message vào cả console và file"""
        print(message, end=end)
        self.file.write(str(message) + (end if end == '\n' else ''))
        self.file.flush()
    
    def close(self):
        """Đóng file"""
        self.file.close()


def main():
    # Tạo log file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = f"facet_comparison_log_{timestamp}.txt"
    logger = Logger(log_file)
    
    try:
        logger.log("━" * 70)
        logger.log("🔍 So sánh kết quả Facet giữa 3 Solr Containers")
        logger.log("━" * 70)
        logger.log(f"\n📋 Số documents: {NUM_DOCS}")
        logger.log(f"📋 Source port: {SOURCE_PORT}")
        logger.log(f"📝 Log file: {log_file}")
        logger.log()
    
        # Bước 1: Lấy danh sách document IDs
        logger.log("━" * 70)
        logger.log("Bước 1: Lấy danh sách document IDs")
        logger.log("━" * 70)
        logger.log()
        
        source_container = CONTAINERS[0]  # Dùng container đầu tiên làm source
        ids = get_document_ids(SOURCE_PORT, source_container["core"], NUM_DOCS)
        
        if not ids:
            logger.log("❌ Không lấy được document IDs. Kiểm tra lại Solr containers.")
            sys.exit(1)
        
        logger.log(f"✅ Đã lấy được {len(ids)} document IDs")
        logger.log(f"   Ví dụ IDs: {ids[:5] if len(ids) >= 5 else ids}")
        logger.log()
    
        # Bước 2: Query từng ID trên cả 3 containers
        logger.log("━" * 70)
        logger.log("Bước 2: Query từng ID trên cả 3 containers")
        logger.log("━" * 70)
        logger.log()
        
        results_dict = defaultdict(dict)
        search_text_dict = {}  # Lưu search_text cho mỗi document
        total_queries = len(ids) * len(CONTAINERS)
        current_query = 0
        
        start_time = time.time()
        
        for idx, doc_id in enumerate(ids, 1):
            logger.log(f"📄 Processing document {idx}/{len(ids)}: {doc_id}")
            
            # Lấy search_text từ container đầu tiên
            search_text = get_search_text(SOURCE_PORT, source_container["core"], doc_id)
            search_text_dict[doc_id] = search_text
            
            for container in CONTAINERS:
                current_query += 1
                logger.log(f"   🔍 Querying {container['version']}...", end=" ")
                
                facets, num_found = get_facet_results(
                    container["port"],
                    container["core"],
                    doc_id
                )
                
                results_dict[doc_id][container["version"]] = facets
                
                if num_found == 0:
                    logger.log(f"⚠️  Document không tồn tại")
                else:
                    logger.log(f"✅ {len(facets)} facet terms")
            
            # Hiển thị progress
            if idx % 10 == 0:
                elapsed = time.time() - start_time
                avg_time = elapsed / current_query
                remaining = (total_queries - current_query) * avg_time
                logger.log(f"   ⏱️  Progress: {current_query}/{total_queries} queries ({idx}/{len(ids)} docs)")
                logger.log(f"   ⏱️  Estimated time remaining: {remaining:.1f}s")
            logger.log()
        
        elapsed_time = time.time() - start_time
        logger.log(f"✅ Hoàn thành query {total_queries} queries trong {elapsed_time:.2f} giây")
        logger.log()
    
        # Bước 3: So sánh kết quả
        logger.log("━" * 70)
        logger.log("Bước 3: So sánh kết quả")
        logger.log("━" * 70)
        logger.log()
        
        comparisons = compare_facet_results(results_dict)
        
        for comp in comparisons:
            logger.log(f"📊 So sánh: {comp['container1']} vs {comp['container2']}")
            logger.log(f"   ✅ Giống nhau: {comp['same']} documents ({comp['same']*100/len(ids):.1f}%)")
            logger.log(f"   ❌ Khác nhau: {comp['different']} documents ({comp['different']*100/len(ids):.1f}%)")
            
            if comp['different'] > 0:
                logger.log(f"   📈 Documents chỉ có trong {comp['container1']}: {comp['only_in_1']}")
                logger.log(f"   📈 Documents chỉ có trong {comp['container2']}: {comp['only_in_2']}")
                logger.log(f"   📊 Trung bình số terms khác nhau: {comp['avg_terms_diff']:.2f}")
            logger.log()
    
        # Bước 4: Tìm các documents có sự khác biệt lớn nhất
        logger.log("━" * 70)
        logger.log("Bước 4: Documents có sự khác biệt lớn nhất")
        logger.log("━" * 70)
        logger.log()
        
        diff_docs = []
        for doc_id in results_dict:
            facets1 = results_dict[doc_id].get(CONTAINERS[0]["version"], {})
            facets2 = results_dict[doc_id].get(CONTAINERS[1]["version"], {})
            facets3 = results_dict[doc_id].get(CONTAINERS[2]["version"], {})
            
            # Tính độ khác biệt
            terms1 = set(facets1.keys())
            terms2 = set(facets2.keys())
            terms3 = set(facets3.keys())
            
            all_terms = terms1 | terms2 | terms3
            common_terms = terms1 & terms2 & terms3
            
            diff_score = len(all_terms) - len(common_terms)
            
            if diff_score > 0:
                diff_docs.append({
                    "id": doc_id,
                    "diff_score": diff_score,
                    "terms_count": {
                        CONTAINERS[0]["version"]: len(terms1),
                        CONTAINERS[1]["version"]: len(terms2),
                        CONTAINERS[2]["version"]: len(terms3)
                    },
                    "only_in": {
                        CONTAINERS[0]["version"]: list(terms1 - terms2 - terms3),
                        CONTAINERS[1]["version"]: list(terms2 - terms1 - terms3),
                        CONTAINERS[2]["version"]: list(terms3 - terms1 - terms2)
                    }
                })
        
        # Sắp xếp theo độ khác biệt
        diff_docs.sort(key=lambda x: x["diff_score"], reverse=True)
        
        # Hiển thị top 10 documents có sự khác biệt lớn nhất
        logger.log("Top 10 documents có sự khác biệt lớn nhất:")
        logger.log()
        for idx, doc in enumerate(diff_docs[:10], 1):
            logger.log(f"{idx}. ID: {doc['id']}")
            logger.log(f"   Độ khác biệt: {doc['diff_score']} terms")
            logger.log(f"   Số terms:")
            for version, count in doc['terms_count'].items():
                logger.log(f"      - {version}: {count}")
            
            # Hiển thị các terms chỉ có trong từng container
            for version, terms in doc['only_in'].items():
                if terms:
                    terms_str = ', '.join(terms[:20])
                    if len(terms) > 20:
                        terms_str += f" ... (và {len(terms) - 20} terms khác)"
                    logger.log(f"   Terms chỉ có trong {version}: {terms_str}")
            logger.log()
    
        # Lưu kết quả chi tiết vào file text
        logger.log("━" * 70)
        logger.log("Bước 5: Chi tiết kết quả từng document")
        logger.log("━" * 70)
        logger.log()
        
        # Log chi tiết các documents có khác biệt
        logger.log(f"📋 Danh sách tất cả {len(diff_docs)} documents có sự khác biệt:")
        logger.log()
        for idx, doc in enumerate(diff_docs, 1):
            logger.log(f"{idx}. Document ID: {doc['id']}")
            logger.log(f"   Độ khác biệt: {doc['diff_score']} terms")
            logger.log(f"   Số terms trong mỗi container:")
            for version, count in doc['terms_count'].items():
                logger.log(f"      - {version}: {count} terms")
            
            # Log chi tiết các terms chỉ có trong từng container
            for version, terms in doc['only_in'].items():
                if terms:
                    logger.log(f"   Terms chỉ có trong {version} ({len(terms)} terms):")
                    # Chia thành các dòng để dễ đọc
                    for i in range(0, len(terms), 10):
                        terms_batch = terms[i:i+10]
                        logger.log(f"      {', '.join(terms_batch)}")
            logger.log()
        
        # Lưu kết quả vào file JSON
        json_file = f"facet_comparison_results_{timestamp}.json"
        output_data = {
            "metadata": {
                "num_docs": len(ids),
                "source_port": SOURCE_PORT,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "elapsed_time": elapsed_time
            },
            "comparisons": comparisons,
            "top_differences": diff_docs[:20],  # Top 20
            "all_differences": diff_docs,  # Tất cả documents có khác biệt
            "sample_results": {doc_id: results_dict[doc_id] for doc_id in ids[:10]}  # Mẫu 10 documents đầu
        }
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)
        
        logger.log(f"💾 Kết quả JSON đã được lưu vào: {json_file}")
        logger.log()
        
        # Xuất ra Excel
        excel_file = export_to_excel(results_dict, search_text_dict, ids, timestamp, logger)
        
        # Tóm tắt
        logger.log("━" * 70)
        logger.log("📊 Tóm tắt")
        logger.log("━" * 70)
        logger.log(f"✅ Đã query {len(ids)} documents trên {len(CONTAINERS)} containers")
        logger.log(f"✅ Tổng số queries: {total_queries}")
        logger.log(f"⏱️  Thời gian thực thi: {elapsed_time:.2f} giây")
        logger.log(f"📈 Tốc độ trung bình: {total_queries/elapsed_time:.2f} queries/giây")
        logger.log(f"📝 Log file: {log_file}")
        logger.log(f"📄 JSON file: {json_file}")
        if excel_file:
            logger.log(f"📊 Excel file: {excel_file}")
        logger.log()
        
    finally:
        logger.close()


if __name__ == "__main__":
    main()
